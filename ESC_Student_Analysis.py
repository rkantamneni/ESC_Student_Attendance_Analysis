import openpyxl
wb = openpyxl.load_workbook('ESC_student_analysis.xlsx', data_only=True)
sheets = wb.sheetnames
total_period_student_names = []
total_period_student_names_no_duplicates = []
global_all_students = []
# Terminology Hierarchy
# Period -> Class -> Session

class Student_Analysis_In_A_Class:

	def __init__(self, name, ws, number_of_sessions, starting_row_of_data, total_participant_col, outputRow):
		self.name = name
		self.ws = ws #Picks the class attendance sheet, four are available
		self.number_of_sessions = number_of_sessions
		self.starting_row_of_data = starting_row_of_data
		self.total_participant_col = total_participant_col
		self.outputRow = outputRow

		global global_all_students
		global total_period_student_names
		self.pcol = 11
		self.currentSheet = self.ws
		self.outputSheet = wb[sheets[0]]
		self.particpants_array = []
		self.all_students = []
		self.numOfStudentsInEachClass = []
		self.attended_1 = self.attended_2 = self.attended_3 = self.attended_4 = 0

		for x in range(1, self.number_of_sessions+1): 
			self.count_students_in_session(x)

		for x in range(1, self.number_of_sessions+1): #Outputs each classs numbers to main page
			self.ws = self.outputSheet
			self.ws.cell(row=self.outputRow, column=x+1).value = self.numOfStudentsInEachClass[x-1]

		self.attended_num_of_sessions(self.total_participant_col, self.outputRow)
		self.percentage_decrease()
		
	def count_students_in_session(self, x):
		self.currentSheet = self.ws
		for row in self.ws.iter_rows(min_row=self.starting_row_of_data, min_col= x, max_col=x, max_row=50):
			for cell in row:
				if not isinstance(cell.value, str):
					self.ws.cell(row=cell.row, column=x).value = "Number of Students Attended: " + str(cell.row-self.starting_row_of_data)
					self.numOfStudentsInEachClass.append(int(cell.row-self.starting_row_of_data))
					return
				global_all_students.append(cell.value)
				self.all_students.append(cell.value)

	def percentage_decrease(self):
		self.currentSheet = self.outputSheet
		self.ws.cell(row=self.outputRow, column=self.pcol).value = ((self.numOfStudentsInEachClass[-1]-self.numOfStudentsInEachClass[0])/(self.numOfStudentsInEachClass[0]))


	def attended_num_of_sessions(self, total_participant_col, outputRow):
		self.ws = self.outputSheet

		#Determines total number of partipcants in class over four sessions
		self.particpants_array = list(set(self.all_students))

		for i in range(len(self.particpants_array)): 
			total_period_student_names.insert(0, self.particpants_array[i])

		self.ws.cell(row=self.outputRow, column=total_participant_col).value = len(self.particpants_array)

		for x in self.particpants_array:
			if self.all_students.count(str(x)) == 1:
				self.attended_1+=1 #This number only attended one session
			if self.all_students.count(str(x)) == 2:
				self.attended_2+=1 #This number only attended two sessions
			if self.all_students.count(str(x)) == 3:
				self.attended_3+=1
			if self.all_students.count(str(x)) == 4:
				self.attended_4+=1

		self.ws.cell(row=self.outputRow, column=6).value = self.attended_1
		self.ws.cell(row=self.outputRow, column=7).value = self.attended_2
		self.ws.cell(row=self.outputRow, column=8).value = self.attended_3
		self.ws.cell(row=self.outputRow, column=9).value = self.attended_4

class total_period_analysis:

	def __init__(self, ws):
		global total_period_student_names_no_duplicates
		self.ws = ws
		self.output_col = 2
		self.partipcateIn1Class = self.partipcateIn2Class= self.partipcateIn3Class= self.partipcateIn4Class = 0
		self.three_or_more_of_2_classes = []
		self.three_or_more_of_3_classes = []
		self.three_or_more_of_4_classes = []
		self.particpatedInNumOfClass()
		self.advancedAnalysis()

	def particpatedInNumOfClass(self):
		total_period_student_names_no_duplicates = list(set(total_period_student_names)) #Takes nams from across classes and removes duplicates

		for x in total_period_student_names_no_duplicates:
			if total_period_student_names.count(str(x)) == 1:
				self.partipcateIn1Class+=1 
			if total_period_student_names.count(str(x)) == 2:
				self.partipcateIn2Class+=1 
			if total_period_student_names.count(str(x)) == 3:
				self.partipcateIn3Class+=1
			if total_period_student_names.count(str(x)) == 4:
				self.partipcateIn4Class+=1

		self.ws.cell(row=8, column=self.output_col).value = self.partipcateIn1Class
		self.ws.cell(row=9, column=self.output_col).value = self.partipcateIn2Class
		self.ws.cell(row=10, column=self.output_col).value = self.partipcateIn3Class
		self.ws.cell(row=11, column=self.output_col).value = self.partipcateIn4Class

	def advancedAnalysis(self):

		total_period_student_names_no_duplicates = list(set(total_period_student_names))
		for x in total_period_student_names_no_duplicates:
			if global_all_students.count(str(x)) >= 3:
				if total_period_student_names.count(str(x)) == 2:
					self.three_or_more_of_2_classes.append(x) 
				if total_period_student_names.count(str(x)) == 3:
					self.three_or_more_of_3_classes.append(x)
				if total_period_student_names.count(str(x)) == 4:
					self.three_or_more_of_4_classes.append(x)

		self.ws.cell(row=12, column=self.output_col).value = (', '.join(self.three_or_more_of_2_classes))
		self.ws.cell(row=13, column=self.output_col).value = (', '.join(self.three_or_more_of_3_classes))
		self.ws.cell(row=14, column=self.output_col).value = (', '.join(self.three_or_more_of_4_classes))



class1 = Student_Analysis_In_A_Class("Python", wb[sheets[1]], 4, 2, 10, 2)

class2 = Student_Analysis_In_A_Class("Sociological Gaze", wb[sheets[2]], 4, 2, 10, 3)

class3 = Student_Analysis_In_A_Class("Portfolio", wb[sheets[3]], 4, 2, 10, 4)

class4 = Student_Analysis_In_A_Class("Tempest", wb[sheets[4]], 4, 2, 10, 5)

overallAnalysis = total_period_analysis(wb[sheets[0]])

wb.save('ESC_student_analysis_complete.xlsx')


		






